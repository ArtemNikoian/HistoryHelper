import re
from openai import OpenAI
import openpyxl
from datetime import datetime
import os 

file_path="path_to_your_excel_file"
api_key="your_openai_api_key"

def get_response(prompt):
    client = OpenAI(api_key=api_key)

    completion = client.chat.completions.create(
    model="gpt-3.5-turbo",
    messages=[
        {"role": "system", "content": "In this chat if I input the Historical event, I want you to solely output the date of the event strictly in the format of xx.yy.zzzz or xx.yy.zzzz - qq.ww.eeee, what people were involved in the event, which countries were involved or where it happened, what is the event, and why it happened. For what and why questions you must write one paragraph of a text for each. Write the text with as simple words as possible, while still being specific. When there are multiple dates, such as for the Hiroshima and Nagasaki bombings, please write them in the following format: xx.yy.zzzz - qq.ww.eeee, where xx.yy.zzzz is the first date and qq.ww.eeee is the second date. If there are more than two dates, use the same format, but write the first date - last date. If you do not know the specific date and only know the year of the event, write it in the format 01.01.xxxx. The date should always be in one of two formats: xx.yy.zzzz or xx.yy.zzzz - qq.ww.eeee. If you are unsure of the specific year, write it as a range of possible dates, like xx.yy.zzzz - qq.ww.eeee. Always remember: the date should always be in the format of xx.yy.zzzz or xx.yy.zzzz - qq.ww.eeee. Perfect examples are: Anschluss\nYour Output:\nDate: 12.03.1938\nPeople Involved: Adolf Hitler, Kurt Schuschnigg\nCountries Involved: Germany, Austria\nEvent: The annexation of Austria by Nazi Germany. It was a major step in Hitler's plan to create a Greater German Reich. He pressured the Austrian government to appoint a pro-German chancellor, and he threatened military action if Austria resisted\nWhy it happened: The Anschluss had deep historical and political roots. Following World War I, the Treaty of Versailles prohibited the union of Germany and Austria. However, the idea of a Greater Germany, incorporating all German-speaking people, gained popularity in Nazi ideology. Hitler, with expansionist goals, aimed to unite all German-speaking regions. The Anschluss also capitalized on the internal political struggles within Austria, allowing Hitler to orchestrate the annexation through both diplomatic pressure and military intervention, solidifying his influence in the region.\n User:\nRhineland remilitarisation\nYour Output:\nDate: 7.03.1936\nPeople Involved: Adolf Hitler\nCountries Involved: Germany, France\nEvent: The remilitarization of the Rhineland, a demilitarized zone in western Germany, was a bold move by Adolf Hitler that violated the Treaty of Versailles, which had ended World War I.\nWhy it happened: Hitler was determined to restore Germany to its former glory and believed that rearming the country was essential to achieving this goal. He also saw the Rhineland as a buffer zone between Germany and France, and he wanted to remove any obstacles to German expansion."},
        {"role": "user", "content": prompt}
    ]
    )
    text=completion.choices[0].message.content
    return text

def find_date():
    pattern = r'Date:\s*(.*?)\n'

    date_match = re.search(pattern, text)
    if date_match:
        date = date_match.group(1)
        return date
    else:
        print("Date not found")
        
def find_people():
    pattern = r'People Involved:\s*(.*?)\n'
    people_match = re.search(pattern, text)
    if people_match:
        people_value = people_match.group(1)
        names = people_value.split(', ')
        return ', '.join(names)
    else:
        print("People Involved not found")
        
def find_countries():
    pattern = r'Countries Involved:\s*(.*?)\n'

    countries_match = re.search(pattern, text)
    if countries_match:
        countries_value = countries_match.group(1)
        country_names = countries_value.split(', ')
        return ', '.join(country_names)
    else:
        print("Countries Involved not found")
        
def find_what():
    pattern = r'Event:\s*(.*?)\n'

    event_match = re.search(pattern, text)
    if event_match:
        event_value = event_match.group(1)
        return str(event_value)
    else:
        print("Event not found")
        
def find_why():
    pattern = r'Why it happened:\s*(.*?)(?:\n|$)'
    why_match = re.search(pattern, text)

    if why_match:
        reason = why_match.group(1)
        return str(reason)
    else:
        print("Why not found")
        
def insert_event(col, event, when, who, where, what, why):
    sheet.cell(row=1, column=col+1).value = event
    sheet.cell(row=2, column=col+1).value = when
    sheet.cell(row=3, column=col+1).value = who
    sheet.cell(row=4, column=col+1).value = where
    sheet.cell(row=5, column=col+1).value = what
    sheet.cell(row=6, column=col+1).value = why

def wrap_text():
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Iterate through all rows and columns
        for row in sheet.iter_rows():
            for cell in row:
                # Wrap text in the cell
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)

def find_init_date(full_date):
    if "-" in full_date:
        init_date, _ = full_date.split(" - ")
        init_date = datetime.strptime(init_date, date_format)
    else:
        init_date = datetime.strptime(full_date, date_format)
    return init_date

while True:
    event_names = input("Write the event names: ").split(", ")
    for event_name in event_names:
        event_done = False
        unsuccessful_trials = 0
        while not event_done or unsuccessful_trials >= 20:
            try:
                text = get_response(event_name)
                event_date = find_date()
                event_people = find_people()
                event_countries = find_countries()
                event_what = find_what()
                event_why = find_why()

                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                date_format = "%d.%m.%Y"
                date_to_fit = find_init_date(event_date)
                col = 2

                while True:
                    start_year_col = str(sheet.cell(row=2, column=col).value)
                    date1 = find_init_date(start_year_col)
                    end_year_col = str(sheet.cell(row=2, column=col + 1).value)
                    date2 = find_init_date(end_year_col)

                    if date1 < date_to_fit < date2 or date1 == date_to_fit:
                        sheet.insert_cols(col + 1, amount=1)
                        insert_event(col, event_name, event_date, event_people, event_countries, event_what, event_why)
                        break
                    elif date_to_fit < date1:
                        sheet.insert_cols(col, amount=1)
                        insert_event(col - 1, event_name, event_date, event_people, event_countries, event_what, event_why)
                        break
                    elif date_to_fit > date2:
                        # Check if cells after date2 are empty
                        empty_cells = True
                        for column in sheet.iter_cols(min_row=2, max_row=2, min_col=col + 2):
                            for cell in column:
                                if cell.value is not None:
                                    empty_cells = False
                                    break
                            if not empty_cells:
                                break
                        if empty_cells:
                            sheet.insert_cols(col + 2, amount=1)
                            insert_event(col+1, event_name, event_date, event_people, event_countries, event_what, event_why)
                            break
                    col += 1
                wrap_text()
                workbook.save(file_path)
                workbook.close()
                event_done = True
                print(f"Event {event_name} done")
            except Exception as e:
                unsuccessful_trials += 1
                print(f"{unsuccessful_trials}th unsuccessful trial for event {event_name}. Wait for more attempts or try a different event.")
                print(f"Error: {e}")
        if unsuccessful_trials >= 20:
            print(f"Too many unsuccessful trials for event {event_name}. There may be an issue with a date of the event.")


